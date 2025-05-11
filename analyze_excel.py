#!/usr/bin/env python3
import pandas as pd
from openpyxl import load_workbook
import sys

def analyze_excel(file_path):
    print(f"Analyzing Excel file: {file_path}\n")
    
    # Get all sheet names
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print(f"Workbook contains {len(sheet_names)} sheets: {', '.join(sheet_names)}\n")
    
    # Analyze each sheet
    for sheet_name in sheet_names:
        try:
            print(f"='SHEET: {sheet_name}'=".center(80, '='))
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Skip entirely empty sheets
            if df.empty or df.columns.empty:
                print("Sheet is empty")
                continue
                
            row_count = len(df)
            col_count = len(df.columns)
            print(f"Rows: {row_count}, Columns: {col_count}")
            
            # Display column headers
            print("\nColumn Headers:")
            for i, col in enumerate(df.columns):
                print(f"  {i+1}: {col}")
            
            # Show sample data
            if row_count > 0:
                print("\nSample Data (first 2 rows):")
                sample = df.head(2)
                for _, row in sample.iterrows():
                    print("  ", end="")
                    for col in df.columns[:10]:  # Show first 10 columns
                        val = row[col]
                        # Truncate long values
                        if isinstance(val, str) and len(val) > 20:
                            val = val[:20] + "..."
                        print(f"{col}: {val}, ", end="")
                    if len(df.columns) > 10:
                        print(f"... ({len(df.columns) - 10} more columns)")
                    else:
                        print()
            
            # Count non-null values
            non_null_counts = df.count()
            print("\nNon-null value counts per column:")
            for col, count in non_null_counts.items():
                if count > 0:  # Only show columns with data
                    print(f"  {col}: {count}")
            
            # Special analysis for documents and contacts sheets
            if "document" in sheet_name.lower() or "summons" in sheet_name.lower() or "aff" in sheet_name.lower():
                print("\n[DOCUMENT ANALYSIS]")
                print(f"This sheet contains {row_count} potential document records")
                
                # Check for duplicate identifiers
                if 'file id' in [col.lower() for col in df.columns]:
                    file_id_col = next(col for col in df.columns if col.lower() == 'file id')
                    unique_file_ids = df[file_id_col].nunique()
                    print(f"  Unique file IDs: {unique_file_ids}")
                    
            if "pm info" in sheet_name.lower() or "client" in sheet_name.lower() or "contact" in sheet_name.lower():
                print("\n[CONTACT ANALYSIS]")
                print(f"This sheet contains {row_count} potential contact records")
                
            print("\n")
        except Exception as e:
            print(f"Error analyzing sheet '{sheet_name}': {e}\n")
    
    print("Analysis complete!")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        analyze_excel(sys.argv[1])
    else:
        print("Please provide an Excel file path as argument")